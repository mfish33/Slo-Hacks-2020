from flask import Flask, send_from_directory 
from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills


def download_workbook(request):
    print(request)
    if request.method == 'OPTIONS':
        # Allows GET requests from any origin with the Content-Type
        # header and caches preflight response for an 3600s
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET, POST',
            'Access-Control-Allow-Headers': 'Content-Type',
            'Access-Control-Max-Age': '3600'
        }
        return ('', 204, headers)
    wb = excel_gen(request.get_json())
    wb.save('/tmp/sample.xlsx')
    r = make_response(send_from_directory('/tmp', 'sample.xlsx'))
    r.mimetype='application/vnd.ms.excel'
    r.headers['Access-Control-Allow-Origin'] = '*'

    return r

def gen_table_styles(ws1, rn, cn, rl, cl, border_type):
    row_num = rn
    col_num = cn
    # location of the Table
    row_loc = rl - 1
    col_loc = cl - 1

    # # Number of Tables
    # Table_num = 1
    # dis_from_expenses = 6  # distance between the tables

    for i in range(row_loc, row_loc + row_num):
        for j in range(col_loc, col_num + col_loc):
            ws1.cell(row=i + 1, column=j + 1).border = border_type

        # row_loc = row_loc + row_num + dis

def gen_fill(ws1, rn, cn, rl, cl, type_fill):
    row_num = rn + 1
    col_num = cn + 1
    # location of the Table
    row_loc = rl - 1 + 1
    col_loc = cl - 1 + 1

    # # Number of Tables
    # Table_num = 1
    # dis_from_expenses = 6  # distance between the tables

    for i in range(row_loc - 1, row_loc + row_num - 2):
        for j in range(col_loc - 1, col_num + col_loc - 2):
            ws1.cell(row=i + 1, column=j + 1).fill = type_fill
        # row_loc = row_loc + row_num + dis


def excel_gen(obj):

    # x = '{"personalInfo":{"sheetName": "Test_Sheet", "income": 100000,"age":22,"state":"California"}, "taxes": {"effectiveTaxRate": 12.3, " totalIncomeTax: 10000, incomeAfterTaxes: 85000"}, "expenses": [{"expense"}]} '
    #x = obj
    # x = """{"PartialSheetInfo":{"investment":{"401k":0,"401kContrib":100,"IRAAmount":0,"IRAType":"","stocks":0}},"expenses":[{"expense":"gas","monthly":250,"weekly":62.5,"yearly":3000},{"expense":"car Payment","monthly":50,"weekly":12.5,"yearly":600},{"expense":"Rent","monthly":900,"weekly":225,"yearly":10800},{"expense":"Food","monthly":500,"weekly":125,"yearly":6000}],"investment":{"401k":2000,"401kContrib":100,"IRAAmount":100,"IRAType":"Roth","stocks":4000},"lifeChoices":{},"personalInfo":{"age":25,"income":150000,"sheetName":"My Financial Plan","state":"California"},"taxes":{"effectiveTaxRate":0.20116333333333333,"incomeAfterTaxes":119825.5,"totalIncomeTax":30174.5}}
    #     """
    j_dic = obj

    thin_border = Border(left=Side(border_style='dashed', color='FF000000'),
                         right=Side(border_style='dashed', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
                         )

    thick_border = Border(left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin', color='FF000000'),
                          top=Side(border_style='thin', color='FF000000'),
                          bottom=Side(border_style='medium', color='FF000000')
                          )
    Double_border = Border(left=Side(border_style='dashed', color='FF000000'),
                           right=Side(border_style='dashed', color='FF000000'),
                           top=Side(border_style='double', color='FF000000'),
                           bottom=Side(border_style='double', color='FF000000')
                           )

    no_border = Border(left=Side(border_style=None, color='FF000000'),
                       right=Side(border_style=None, color='FF000000'),
                       top=Side(border_style=None, color='FF000000'),
                       bottom=Side(border_style=None, color='FF000000')
                       )

    border_list = [thin_border, thick_border, Double_border, no_border]

    # Define fill formating
    yellow = PatternFill(fill_type=fills.FILL_SOLID, start_color='00FFD166', end_color='00FFFF00')
    green = PatternFill(fill_type=fills.FILL_SOLID, start_color='0099BA66', end_color='00FFFF00')
    turq = PatternFill(fill_type=fills.FILL_SOLID, start_color='0013A192', end_color='00FFFF00')
    blue = PatternFill(fill_type=fills.FILL_SOLID, start_color='0079A4EF', end_color='00FFFF00')
    grey = PatternFill(fill_type=fills.FILL_SOLID, start_color='00516C6E', end_color='00FFFF00')
    fill_list = []
    fill_list.append(blue)
    fill_list.append(green)
    fill_list.append(grey)
    fill_list.append(yellow)
    fill_list.append(turq)

    title_font = Font(name='Calibri', color=colors.BLACK, bold=True, size=14, underline='none')
    ft_1 = Font(name='Calibri', color=colors.BLUE, bold=True, size=11, underline='none')
    ft_2 = Font(name='Calibri', color=colors.GREEN, bold=True, size=11, underline='none')
    ft_3 = Font(name='Calibri', color=colors.RED, bold=True, size=11, underline='none')
    ft_4 = Font(name='Calibri', color=colors.BLACK, bold=True, size=11, underline='none')

    align = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True,
                      shrink_to_fit=True, indent=0)

    wb = Workbook()
    ws1 = wb.active  # work sheet
    ws1.title = j_dic["personalInfo"]["sheetName"]
    ws1.sheet_properties.tabColor = '005094BD'

    # Personal Info
    age = j_dic["personalInfo"]["age"]
    income = j_dic["personalInfo"]["income"]
    state = j_dic["personalInfo"]["state"]
    age_r = 2
    age_c = 2
    ws1.merge_cells(start_row=age_r, start_column=age_c, end_row=age_r, end_column=age_c + 1)
    ws1[get_column_letter(age_c) + str(age_r)] = "Personal Information"
    ws1[get_column_letter(age_c) + str(age_r)].alignment = align
    ws1[get_column_letter(age_c) + str(age_r)].font = title_font
    ws1[get_column_letter(age_c) + str(age_r+1)] = "Age"
    ws1[get_column_letter(age_c) + str(age_r+1)].alignment = align
    ws1[get_column_letter(age_c + 1) + str(age_r+1)] = age
    ws1[get_column_letter(age_c + 1) + str(age_r + 1)].alignment = align
    ws1[get_column_letter(age_c) + str(age_r + 2)] = "Income"
    ws1[get_column_letter(age_c) + str(age_r+2)].alignment = align
    ws1[get_column_letter(age_c + 1) + str(age_r+2)] = income
    ws1[get_column_letter(age_c + 1) + str(age_r + 2)].alignment = align
    ws1[get_column_letter(age_c) + str(age_r + 3)] = "State"
    ws1[get_column_letter(age_c) + str(age_r+3)].alignment = align
    ws1[get_column_letter(age_c + 1) + str(age_r+3)] = state
    ws1[get_column_letter(age_c + 1) + str(age_r + 3)].alignment = align
    gen_fill(ws1, 1, 2, age_r, age_c, fill_list[1])
    gen_fill(ws1, 3, 1, age_r+1, age_c, fill_list[0])
    gen_fill(ws1, 3, 1, age_r + 1, age_c+1, fill_list[3])
    gen_table_styles(ws1, 4, 2, age_r, age_c, border_list[1])

    #Taxes
    taxes_r = 7
    taxes_c = 2
    eftaxrate = j_dic["taxes"]["effectiveTaxRate"]
    tot_inc_tax = j_dic["taxes"]["totalIncomeTax"]
    inc_a_tax = j_dic["taxes"]["incomeAfterTaxes"]
    ws1.merge_cells(start_row=taxes_r, start_column=taxes_c, end_row=taxes_r, end_column=taxes_c + 1)
    ws1[get_column_letter(taxes_c) + str(taxes_r)] = "Taxes"
    ws1[get_column_letter(taxes_c) + str(taxes_r)].alignment = align
    ws1[get_column_letter(taxes_c) + str(taxes_r)].font = title_font
    ws1[get_column_letter(taxes_c) + str(taxes_r + 1)] = "Effective Tax Rate"
    ws1[get_column_letter(taxes_c) + str(taxes_r + 1)].alignment = align
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 1)] = eftaxrate
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 1)].alignment = align
    ws1[get_column_letter(taxes_c) + str(taxes_r + 2)] = "Total Income Tax"
    ws1[get_column_letter(taxes_c) + str(taxes_r + 2)].alignment = align
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 2)] = tot_inc_tax
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 2)].alignment = align
    ws1[get_column_letter(taxes_c) + str(taxes_r + 3)] = "Income After Taxes"
    ws1[get_column_letter(taxes_c) + str(taxes_r + 3)].alignment = align
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 3)] = inc_a_tax
    ws1[get_column_letter(taxes_c + 1) + str(taxes_r + 3)].alignment = align
    gen_fill(ws1, 1, 2, taxes_r, taxes_c, fill_list[1])
    gen_fill(ws1, 3, 1, taxes_r + 1, taxes_c, fill_list[0])
    gen_fill(ws1, 3, 1, taxes_r + 1, taxes_c + 1, fill_list[3])
    gen_table_styles(ws1, 4, 2, taxes_r, taxes_c, border_list[1])

    #investments
    inv_r = 2
    inv_c = 5

    fourk = j_dic["investment"]["401k"]
    contrib = j_dic["investment"]["401kContrib"]
    iraamount = j_dic["investment"]["IRAAmount"]
    ira_type = j_dic["investment"]["IRAType"]
    stocks = j_dic["investment"]["stocks"]

    ws1.merge_cells(start_row=inv_r, start_column=inv_c, end_row=inv_r, end_column=inv_c + 1)
    ws1[get_column_letter(inv_c) + str(inv_r)] = "Investment"
    ws1[get_column_letter(inv_c) + str(inv_r)].alignment = align
    ws1[get_column_letter(inv_c) + str(inv_r)].font = title_font
    ws1[get_column_letter(inv_c) + str(inv_r + 1)] = "401k"
    ws1[get_column_letter(inv_c) + str(inv_r + 1)].alignment = align
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 1)] = fourk
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 1)].alignment = align
    ws1[get_column_letter(inv_c) + str(inv_r + 2)] = "401k Contributions"
    ws1[get_column_letter(inv_c) + str(inv_r + 2)].alignment = align
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 2)] = contrib
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 2)].alignment = align
    ws1[get_column_letter(inv_c) + str(inv_r + 3)] = "IRA Amount"
    ws1[get_column_letter(inv_c) + str(inv_r + 3)].alignment = align
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 3)] = iraamount
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 3)].alignment = align
    ws1[get_column_letter(inv_c) + str(inv_r + 4)] = "IRA Type"
    ws1[get_column_letter(inv_c) + str(inv_r + 4)].alignment = align
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 4)] = ira_type
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 4)].alignment = align
    ws1[get_column_letter(inv_c) + str(inv_r + 5)] = "Stocks"
    ws1[get_column_letter(inv_c) + str(inv_r + 5)].alignment = align
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 5)] = stocks
    ws1[get_column_letter(inv_c + 1) + str(inv_r + 5)].alignment = align
    gen_fill(ws1, 1, 2, inv_r, inv_c, fill_list[1])
    gen_fill(ws1, 5, 1, inv_r + 1, inv_c, fill_list[0])
    gen_fill(ws1, 5, 1, inv_r + 1, inv_c + 1, fill_list[3])
    gen_table_styles(ws1, 6, 2, inv_r, inv_c, border_list[1])

    # Expenses Section
    # listA = {"expense": "Gas", "weekly": 50.0, "monthly": 217.86, "yearly": 2600.0}
    # listB = {"expense": "Food", "weekly": 200.0, "monthly": 817.43, "yearly": 10400.0}
    # listC = {"expense": "Car M", "weekly": 50.0, "monthly": 200.00, "yearly": 2400.0}
    # d_of_exp = [listA, listB, listC]  # Form->expenses:expense[]
    l_of_exp = []  # matrix of all the expenses
    for hmap in j_dic["expenses"]:
        temp = []
        for k in hmap.keys():
            temp.append(hmap[k])
        l_of_exp.append(temp)


    # Expenses Tracker
    posttax_r = 2
    posttax_c = 8
    ws1.merge_cells(start_row=posttax_r, start_column=posttax_c+1, end_row=posttax_r, end_column=posttax_c + 4)
    ws1[get_column_letter(posttax_c+1) + str(posttax_r)] = "Expenses Tracker"
    ws1[get_column_letter(posttax_c+1) + str(posttax_r)].alignment = align
    ws1[get_column_letter(posttax_c+1) + str(posttax_r)].font = title_font
    ws1[get_column_letter(posttax_c + 1) + str(posttax_r + 1)] = "Expenses"
    ws1[get_column_letter(posttax_c + 1) + str(posttax_r + 1)].alignment = align
    ws1[get_column_letter(posttax_c + 2) + str(posttax_r + 1)] = "Weekly"
    ws1[get_column_letter(posttax_c + 2) + str(posttax_r + 1)].alignment = align
    ws1[get_column_letter(posttax_c + 3) + str(posttax_r + 1)] = "Monthly"
    ws1[get_column_letter(posttax_c + 3) + str(posttax_r + 1)].alignment = align
    ws1[get_column_letter(posttax_c + 4) + str(posttax_r + 1)] = "Yearly"
    ws1[get_column_letter(posttax_c + 4) + str(posttax_r + 1)].alignment = align
    # ws1[get_column_letter(posttax_c + 5) + str(posttax_r + 1)] = "Percent of Post Tax Income"
    # ws1[get_column_letter(posttax_c + 5) + str(posttax_r + 1)].alignment = align

    ws1.merge_cells(start_row=posttax_r + 2, start_column=posttax_c, end_row=posttax_r + len(l_of_exp) + 1,
                    end_column=posttax_c)
    ws1[get_column_letter(posttax_c) + str(posttax_r + 2)] = "Living Expenses"
    ws1[get_column_letter(posttax_c) + str(posttax_r + 2)].alignment = align

    # ws1.merge_cells(start_row=posttax_r + len(l_of_exp) + 2, start_column=posttax_c,
    #                 end_row=posttax_r + len(l_of_exp) + 4,
    #                 end_column=posttax_c)
    # ws1[get_column_letter(posttax_c) + str(posttax_r + len(l_of_exp) + 2)] = "Post Tax Investing"
    # ws1[get_column_letter(posttax_c) + str(posttax_r + len(l_of_exp) + 2)].alignment = align

    # gen_fill(ws1, 2, 6, posttax_r, posttax_c, fill_list[1])
    # gen_fill(ws1, len(l_of_exp), 2, posttax_r + 2, posttax_c, fill_list[0])
    # gen_fill(ws1, 3, 2, posttax_r + len(l_of_exp) + 2, posttax_c, fill_list[3])
    # gen_table_styles(ws1, 5 + len(l_of_exp), 6, posttax_r, posttax_c, border_list[1])

    for row in range(len(l_of_exp)):  # because in the expenses there are only four values
        for col in range(4):
            _ = ws1.cell(column=col + posttax_c + 1, row=row + posttax_r + 2, value=l_of_exp[row][col])
            ws1.cell(column=col + posttax_c + 1, row=row + posttax_r + 2, value=l_of_exp[row][col]).alignment = align
            # if col == 3:
            #     _ = ws1.cell(column=col + posttax_c + 2, row=row + posttax_r + 2,
            #                  value='=(' + get_column_letter(col + posttax_c + 3) + str(
            #                      row + posttax_r + 3) + '/' + get_column_letter(
            #                      col + posttax_c + 3) + str(row + posttax_r + 4) + ')')
            #     _.font=ft_2

    gen_fill(ws1, 2, 5, posttax_r, posttax_c, fill_list[1])
    # gen_fill(ws1, 5, 1, posttax_r +2, posttax_c, fill_list[0])
    gen_fill(ws1, len(j_dic["expenses"]), 2, posttax_r +2, posttax_c, fill_list[0])
    gen_table_styles(ws1, len(j_dic["expenses"])+2, 5, posttax_r, posttax_c, border_list[1])

    gen_fill(ws1, len(j_dic["expenses"]), 4, posttax_r+2, posttax_c+1, fill_list[3])
    gen_table_styles(ws1, len(j_dic["expenses"]) + 2, 5, posttax_r, posttax_c, border_list[0])



    # after_tax_col_ref = 8  # 8 because in sample sheet it corresponds to "After Taxes" being in col H

    # # Expenses Totals
    # ws1[get_column_letter(after_tax_col_ref) + str(len(l_of_exp) + 8)] = 'After Taxes'
    # ws1[get_column_letter(after_tax_col_ref + 1) + str(len(l_of_exp) + 7)] = 'Weekly Total'
    # ws1[get_column_letter(after_tax_col_ref + 2) + str(len(l_of_exp) + 7)] = 'Monthly Total'
    # ws1[get_column_letter(after_tax_col_ref + 3) + str(len(l_of_exp) + 7)] = 'Yearly Total'

    return wb


